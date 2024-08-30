from openpyxl import load_workbook
from django.conf import settings
import pandas as pd
import os 

def update_userdata(data):
    """
    주어진 데이터를 userdata.xlsx 파일에 적용하여 기존 데이터를 수정하는 함수.
    
    :param data: 리스트 형식의 사용자 정보 예시: ['승환', '202134232', '컴공학과', '컴공전공', '2학년', '휴학', '2021', '일반학사']
    """
    file_path = settings.BASE_DIR / 'myapp' / 'data' / 'userdata.xlsx'  # 엑셀 파일 경로 설정

    # 엑셀 파일 불러오기
    workbook = load_workbook(file_path)
    sheet = workbook.active

    for col, value in enumerate(data, start=1):
    # col은 숫자이므로 이를 string으로 변환할 필요 없음, value는 문자열로 유지
        sheet.cell(row=2, column=col).value = str(value)  # 필요에 따라 str()로 변환
    # 변경된 내용을 저장
    workbook.save(file_path)
    workbook.close()

def update_course_row(course, selected_course):
    """
    주어진 과목(course) 파일에서 선택된 과목명(selected_course)을 찾아 해당 행을 반환하고,
    그 행을 원본 파일에서 삭제한 후 업데이트합니다.
    """
    file_path = os.path.join(settings.BASE_DIR, 'myapp', 'data', f'{course}.xlsx')
    if os.path.exists(file_path):
        df = pd.read_excel(file_path)
        # 선택된 과목명에 해당하는 행을 찾습니다.
        row_to_move = df[df['이수현황'] == selected_course]
        if not row_to_move.empty:
            # 해당 행을 제거한 데이터프레임으로 업데이트
            df = df[df['이수현황'] != selected_course]
            df.to_excel(file_path, index=False)  # 원본 파일 업데이트
            return row_to_move
    return None
def update_row_to_target(row, target_file_name='자유선택.xlsx'):
    """
    주어진 행(row)을 자유선택.xlsx 파일의 가장 아래로 추가합니다.
    """
    target_file_path = os.path.join(settings.BASE_DIR, 'myapp', 'data', target_file_name)
    if os.path.exists(target_file_path):
        target_df = pd.read_excel(target_file_path)
        # 해당 행을 자유선택 파일의 가장 아래로 추가합니다.
        target_df = pd.concat([target_df, row], ignore_index=True)
        # 자유선택 파일을 다시 저장합니다.
        target_df.to_excel(target_file_path, index=False)
    else:
        print(f"Target file {target_file_name} not found.")
def update_checked_courses(checked_courses, dropdown_values): #초과학점 선택된 것을 자유학점으로 넘기는 함수 void

    """
    체크된 과목들에 대해 자유선택.xlsx 파일로 행을 이동시키고 원본 파일에서 제거하는 전체 과정을 처리합니다.
    """
    for course in checked_courses:
        selected_course = dropdown_values.get(course)
        row_to_move = update_course_row(course, selected_course)
        if row_to_move is not None:
            update_row_to_target(row_to_move)
